# ==============================================
# CPENN DATA & PROJECTIONS
# NFL + NASCAR + MLB Projections Explorer (multi-sport)
# ==============================================

import os
import re
import json
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# ----------------------------
# STREAMLIT PAGE CONFIGURATION
# ----------------------------
st.set_page_config(
    page_title="CPENN DATA & PROJECTIONS",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Small global CSS touch-ups
st.markdown("""
<style>
/* tighten paddings */
.block-container { padding-top: 1.1rem; padding-bottom: 1.1rem; }
/* smaller element gaps */
.element-container { margin-bottom: 0.6rem; }
/* hide dataframe row index */
[data-testid="stDataFrame"] div[role="gridcell"][data-testid="stRowHeaderCell"] { display: none; }
/* nicer tabs */
.stTabs [data-baseweb="tab-list"] { gap: 0.5rem; }
.stTabs [data-baseweb="tab"] { background: #0f172a12; border-radius: 10px; padding: 0.5rem 0.9rem; }
</style>
""", unsafe_allow_html=True)

# ----------------------------
# APP HEADER
# ----------------------------
st.markdown(
    """
<div style='text-align:center; padding: 1rem; background: linear-gradient(90deg, #111827 0%, #1f2937 60%, #0b1220 100%); border-radius: 12px; margin-bottom: 0.9rem; border: 1px solid #1f2937;'>
  <h1 style='color: #e5e7eb; margin: 0; letter-spacing: 0.6px;'>CPENN DATA & PROJECTIONS</h1>
  <p style='color: #9ca3af; margin: 0.2rem 0 0; font-size: 0.95rem;'>Clean • Fast • Uniform Across DK & FD</p>
</div>
""",
    unsafe_allow_html=True,
)

# ----------------------------
# DEFAULT LOCAL DATA SOURCES
# ----------------------------
DEFAULT_SPORTS = {
    "NFL": [
        {
            "label": "NFL Week 1 Main",
            "path": r"C:\Users\cpenn\Dropbox\Sports Models\NFL\NFL Week 1 Main.xlsm",
            "sheets": ["Projections", "Stacks", "QB Projections", "RB Projections", "WR Projections", "TE Projections"],
        },
        {
            "label": "Superbowl Showdown",
            "path": r"C:\Users\cpenn\Dropbox\Sports Models\NFL\NFL Superbowl Showdown Chiefs vs Eagles.xlsm",
            "sheets": ["Projections", "QB Projections", "RB Projections", "WR Projections", "TE Projections"],
        },
    ],
    "NASCAR": [
        {
            "label": "Cup – Daytona",
            "path": r"C:\Users\cpenn\Dropbox\Sports Models\2025 NASCAR\Cup Coke 400 Daytona.xlsm",
            "sheets": ["Projections", "Betting Dashboard"],
        },
        {
            "label": "Trucks – Richmond",
            "path": r"C:\Users\cpenn\Dropbox\Sports Models\2025 NASCAR\Trucks Eero 250 Richmond.xlsm",
            "sheets": ["Projections", "Betting Dashboard"],
        }
    ],
    # ----------------------------
    # MLB (NEW)
    # ----------------------------
    "MLB": [
        {
            "label": "MLB – Daily",
            "path": r"C:\Users\cpenn\Dropbox\Sports Models\MLB\MLB August 15th.xlsm",
            "sheets": ["Pitcher Projections", "Batter Projections", "Top Stacks"],
        }
    ],
}

# ----------------------------
# UI logging (silencable)
# ----------------------------
def ui_log(msg: str, level: str = "info"):
    """Respect the global 'Show load logs' switch in sidebar."""
    if not st.session_state.get("show_logs", False):
        return
    getattr(st, level, st.info)(msg)

# ----------------------------
# NASCAR TARGET COLUMN WHITELISTS
# ----------------------------
DESIRED_NASCAR_PROJECTIONS = [
    "Driver",
    "Qual",
    "DK Sal", "FD Sal",
    "DK Proj", "DK Val", "FD Proj", "FD Val",
    "Proj Fin",
    "pLL", "pFL",
    "DK PP", "DK Dom", "FD PP", "FD Dom",
    "DK Floor", "DK Ceiling", "FD Floor", "FD Ceiling",
    "DK pOWN%", "FD pOWN%",
    "DK Opt%", "FD Opt%",
    "DK Lev%", "FD Lev%",
    "DK Rtg", "FD Rtg"
]
DESIRED_NASCAR_BETTING = [
    "Driver", "Proj Fin", "Win%", "T3%", "T5%", "T10%", "Win", "T3", "T5", "T10"
]

# ----------------------------
# MLB TARGET COLUMN WHITELISTS (UPDATED to your headers)
# ----------------------------
DESIRED_MLB_PITCHERS = [
    "V", "H", "Player", "DK Sal", "FD Sal", "Team", "Opp",
    "IP", "ER", "K", "H", "BB", "HR", "W",
    "DK Proj", "DK Val", "DK pOWN",
    "FD Proj", "FD Val", "FD pOWN",
    "DK F", "DK C", "FD F", "FD C",
    "DK Rtg", "FD Rtg",
]

DESIRED_MLB_BATTERS = [
    "BO", "H", "Pos", "V", "Player", "DK Sal", "FD Sal", "Team", "Opp",
    "AB", "H", "1B", "2B", "3B", "HR", "RBI", "R", "SB", "BB", "K",
    "DK Proj", "DK Val", "DK pOWN",
    "FD Proj", "FD Val", "FD pOWN",
    "DK F", "DK C", "FD F", "FD C",
    "DK Rtg", "FD Rtg",
]

# Stacks sheet (kept flexible but aligned to your short labels)
DESIRED_MLB_STACKS = [
    "Team", "DK Sal", "FD Sal", "Total", "Park", "Opp", "Opp Pitcher", "H",
    "DK Proj", "Val", "Own%", "Tstack%", "vStack%", "Lev%",
    "FD Proj", "FD Val", "FD Own%", "FD Tstack%", "FD vStack%", "FD Lev%",
]

# ----------------------------
# HEADER / ALIAS HANDLING
# ----------------------------
def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())

_NASCAR_ALIAS = {
    "Driver":      ["Driver","driver","Player Name","Player","Name"],
    "Qual":        ["Qual","Start Pos","Start","Starting","Qualifying"],
    "Proj Fin":    ["Proj Fin","ProjFinish","Proj Finish","Projected Finish","ProjFinish."],
    "pLL":         ["pLL","LL","Place Laps Led","Prob LL"],
    "pFL":         ["pFL","FL","Fast Laps Prob"],
    "DK Sal":      ["DK Sal","DK Salary","DraftKings Salary"],
    "DK Proj":     ["DK Proj","DK Proj.","DK Projection","DK Pro"],
    "DK Val":      ["DK Val","DK Value"],
    "DK pOWN%":    ["DK pOWN%","DK Own%","DK Ownership%","DK Ownership","pOWN% (DK)"],
    "DK Opt%":     ["DK Opt%","DK Optimal%","DK Optimal"],
    "DK Lev%":     ["DK Lev%","DK Leverage%","DK Leverage"],
    "DK Rtg":      ["DK Rtg","DK RTG","DraftKings Rating"],
    "DK PP":       ["DK PP","DK Pts/P$","DK Points per $"],
    "DK Dom":      ["DK Dom","DK Dominator"],
    "DK Floor":    ["DK Floor"],
    "DK Ceiling":  ["DK Ceiling","DK Ceil"],
    "FD Sal":      ["FD Sal","FD Salary","FanDuel Salary"],
    "FD Proj":     ["FD Proj","FD Proj.","FD Projection","FD Pro"],
    "FD Val":      ["FD Val","FD Value"],
    "FD pOWN%":    ["FD pOWN%","FD Own%","FD Ownership%","FD Ownership","pOWN% (FD)"],
    "FD Opt%":     ["FD Opt%","FD Optimal%","FD Optimal"],
    "FD Lev%":     ["FD Lev%","FD Leverage%","FD Leverage"],
    "FD Rtg":      ["FD Rtg","FD RTG","FanDuel Rating"],
    "FD PP":       ["FD PP","FD Pts/P$","FD Points per $"],
    "FD Dom":      ["FD Dom","FD Dominator"],
    "FD Floor":    ["FD Floor"],
    "FD Ceiling":  ["FD Ceiling","FD Ceil","FD Ceilig"],
    "Win%":        ["Win%","Win %","Prob Win","P(Win)","Race Win %"],
    "T3%":         ["T3%","Top 3 %","Prob T3","P(Top 3)","Top3%"],
    "T5%":         ["T5%","Top 5 %","Prob T5","P(Top 5)","Top5%"],
    "T10%":        ["T10%","Top 10 %","Prob T10","P(Top 10)","Top10%"],
    "Win":         ["Win","Winner Odds","Race Winner","Best Win Odds","Win ML","Win Price"],
    "T3":          ["T3","Top 3","Top 3 Odds","Best Top 3 Odds"],
    "T5":          ["T5","Top 5","Top 5 Odds","Best Top 5 Odds"],
    "T10":         ["T10","Top 10","Top 10 Odds","Best Top 10 Odds"],
}

# ----------------------------
# MLB HEADER / ALIAS HANDLING (UPDATED)
# ----------------------------
_MLB_ALIAS = {
    # Canonical short display labels (left) ← acceptable variants (right)

    # Shared / identity
    "Player": ["Player", "player", "Player Name", "Name"],
    "Pos": ["Pos", "Position"],
    "Team": ["Team", "Tm"],
    "Opp": ["Opp", "Opponent"],
    "BO": ["BO", "Bat Order", "Order"],
    "H": ["H"],  # heat/conditional helper or Hits (handled by position of column)

    # Implied total (column is literally 'V' in your files)
    "V": ["V", "Imp Tot", "Imp. Tot", "Implied Total", "Team Imp. Tot", "Team Total", "Vegas Team Total", "TTTL", "IR"],

    # Salaries
    "DK Sal": ["DK Sal", "DK Salary", "DraftKings Salary", "DK Price", "DK$"],
    "FD Sal": ["FD Sal", "FD Salary", "FanDuel Salary", "FD Price", "FD$"],

    # Projections / Values
    "DK Proj": ["DK Proj", "DK Projection", "DK Proj.", "DK Points"],
    "FD Proj": ["FD Proj", "FD Projection", "FD Proj.", "FD Points"],
    "DK Val":  ["DK Val", "DK Value", "Value (DK)"],
    "FD Val":  ["FD Val", "FD Value", "Value (FD)"],

    # Ownership (no % in your short labels)
    "DK pOWN": ["DK pOWN", "DK pOWN%", "DK Own%", "DK Ownership%", "Ownership% (DK)", "pOWN% (DK)"],
    "FD pOWN": ["FD pOWN", "FD pOWN%", "FD Own%", "FD Ownership%", "Ownership% (FD)", "pOWN% (FD)"],

    # Floors/Ceilings (short)
    "DK F": ["DK F", "DK Floor"],
    "DK C": ["DK C", "DK Ceiling", "DK Ceil"],
    "FD F": ["FD F", "FD Floor"],
    "FD C": ["FD C", "FD Ceiling", "FD Ceil"],

    # Ratings
    "DK Rtg": ["DK Rtg", "DK RTG", "DraftKings Rating"],
    "FD Rtg": ["FD Rtg", "FD RTG", "FanDuel Rating"],

    # Pitcher detailed stats
    "IP": ["IP", "IP Proj", "Innings Proj", "Projected IP"],
    "ER": ["ER"],
    "K":  ["K", "K Proj", "K Projection", "Proj K", "Ks Proj"],
    "BB": ["BB"],
    "HR": ["HR", "HR/9"],  # if you carry HR/9 separately, it’ll just be ignored here
    "W":  ["W", "Win%", "P(Win)", "Win Prob"],

    # Batter counting stats
    "AB": ["AB"], "1B": ["1B"], "2B": ["2B"], "3B": ["3B"],
    "RBI": ["RBI"], "R": ["R"], "SB": ["SB"],
}

def _apply_aliases(df: pd.DataFrame, alias_map: Dict[str, List[str]]) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    inv = {}
    for canon, alist in alias_map.items():
        for a in alist:
            inv[_norm(a)] = canon
    ren = {}
    for c in df.columns:
        cn = _norm(c)
        if cn in inv:
            ren[c] = inv[cn]
    if ren:
        df = df.rename(columns=ren)
    return df

# ----------------------------
# GENERIC CLEANUP CONSTANTS
# ----------------------------
NUMERIC_KEYS = {
    "DK Sal", "FD Sal", "DK Proj", "FD Proj", "DK Val", "FD Val",
    "DK pOWN%", "FD pOWN%", "DK Opt%", "FD Opt%", "DK Lev%", "FD Lev%",
    "DK Rtg", "FD Rtg", "DK Pro", "FD Pro",
    "Comp%", "Pa Yards", "Pa Attempts", "Pa Comp", "Pa TD", "Int",
    "Ru Attempts", "Ru Yards", "Ru TD", "Rec", "Rec Yards", "Rec TD",
    "DK%", "FD%",
    "Qual", "Start Pos", "Proj Finish", "Proj Fin", "pLL", "Fast Laps", "PD",
    "DK PP", "FD PP", "DK Dom", "FD Dom", "DK Floor", "DK Ceiling", "FD Floor", "FD Ceilig",
}

EXCLUDE_PATTERNS = [
    re.compile(r"^unnamed", re.I),
    re.compile(r"opto import$", re.I),
    re.compile(r"^_?\d+(?:\.\d+)?$"),
]

POSITION_COLORS = {
    "QB": "#FF6B6B",
    "RB": "#4ECDC4",
    "WR": "#45B7D1",
    "TE": "#96CEB4",
    "K": "#FFEAA7",
    "DST": "#DDA0DD",
}

PRESET_FILE = Path(__file__).with_name("column_presets.json")

# ----------------------------
# MLB rate stats that need 3-dec display (NEW)
# ----------------------------
MLB_THREE_DEC_STATS = {"ba", "avg", "obp", "slg", "ops", "woba", "xwoba"}

# ----------------------------
# PRESET HELPERS
# ----------------------------
def _load_preset_store() -> Dict[str, Dict[str, List[str]]]:
    if PRESET_FILE.exists():
        try:
            return json.loads(PRESET_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def _save_preset_store(store: Dict[str, Dict[str, List[str]]]) -> None:
    try:
        PRESET_FILE.write_text(json.dumps(store, indent=2), encoding="utf-8")
    except Exception:
        pass

# ----------------------------
# GENERIC COLUMN STANDARDIZATION (NFL/general)
# ----------------------------
def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    renames: Dict[str, str] = {}
    cols_lower = {c: str(c).strip().lower() for c in df.columns}
    for c, lc in cols_lower.items():
        if lc in ("player", "name", "playername", "player name"):
            renames[c] = "Player Name"
        elif lc in ("position", "pos."):
            renames[c] = "Pos"
        elif lc in ("opponent", "opp."):
            renames[c] = "Opp"
        elif lc == "teamname":
            renames[c] = "Team"
    if renames:
        df = df.rename(columns=renames)
    return df

def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    keep = [c for c in df.columns if not any(rx.search(str(c)) for rx in EXCLUDE_PATTERNS)]
    df = df[keep].copy()
    df.columns = [str(c).strip() for c in df.columns]
    df = standardize_columns(df)
    for col in df.columns:
        if col in NUMERIC_KEYS:
            try:
                series = df[col].astype(str).str.replace(r"[%$,]", "", regex=True)
                df[col] = pd.to_numeric(series, errors="coerce")
            except Exception:
                pass
    empty_cols = [c for c in df.columns if df[c].isna().all() or (df[c].astype(str).str.strip() == "").all()]
    if empty_cols:
        df = df.drop(columns=empty_cols)
    return df

# ----------------------------
# FAST NASCAR SHEET READER
# ----------------------------
@st.cache_data(show_spinner=False, ttl=60)
def _fast_read_nascar_sheet(path_or_file, sheet_name: str,
                            desired_columns: List[str]) -> Optional[pd.DataFrame]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path_or_file, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]

        header_row_ix, header_vals = None, None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if not row or not any(row):
                continue
            lows = {_norm(str(v)) for v in row if v is not None}
            hints = {"driver","playername","player","projfin","projfinish","win","t3","t5","t10","odds","qual","start","dk","fd"}
            if any(h in "".join(lows) for h in hints):
                header_row_ix, header_vals = r_idx, list(row)
                break
        if header_row_ix is None or not header_vals:
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
                if row and any(v is not None for v in row):
                    header_row_ix, header_vals = r_idx, list(row)
                    break
        if header_row_ix is None or not header_vals:
            return None

        actual_map = {str(v).strip(): i for i, v in enumerate(header_vals) if v is not None and str(v).strip()}
        desired_to_actual, indices = {}, []
        for canonical in desired_columns:
            alias_list = _NASCAR_ALIAS.get(canonical, [canonical])
            want_norms = {_norm(a) for a in alias_list}
            for actual, idx in actual_map.items():
                if _norm(actual) in want_norms:
                    desired_to_actual[canonical] = actual
                    indices.append(idx)
                    break

        if not indices:
            return None

        df = pd.read_excel(
            path_or_file,
            sheet_name=sheet_name,
            engine="openpyxl",
            header=header_row_ix - 1,
            usecols=sorted(set(indices)),
        )

        df = df.rename(columns={v: k for k, v in desired_to_actual.items()})

        if "Driver" not in df.columns:
            for cand in ["Player Name", "Player", "Name"]:
                if cand in df.columns:
                    df = df.rename(columns={cand: "Driver"})
                    break

        # de-dupe headers
        df.columns = pd.Index([f"{col}_{i}" if list(df.columns).count(col) > 1 and list(df.columns)[:i].count(col) > 0 else col 
                              for i, col in enumerate(df.columns)])

        safe_columns = []
        for col in desired_columns:
            if col in df.columns:
                safe_columns.append(col)
            else:
                safe_columns.append(col)
        df = df.reindex(columns=safe_columns)

        for c in df.columns:
            if c == "Driver":
                continue
            df[c] = pd.to_numeric(
                pd.Series(df[c]).astype(str).str.replace(r"[%$,]", "", regex=True),
                errors="coerce"
            )

        if "Qual" in df.columns and "Proj Fin" in df.columns and "PD" not in df.columns:
            try:
                df["PD"] = pd.to_numeric(df["Qual"], errors="coerce") - pd.to_numeric(df["Proj Fin"], errors="coerce")
            except Exception:
                pass

        return df
    except Exception as e:
        ui_log(f"Error reading NASCAR sheet {sheet_name}: {str(e)}", "error")
        return None

# ----------------------------
# SHEET RESOLUTION / RAW READERS
# ----------------------------
def resolve_allowed_sheets(path_or_file, desired: Optional[List[str]]) -> List[str]:
    try:
        xls = pd.ExcelFile(path_or_file, engine="openpyxl")
        actual = xls.sheet_names
        if desired is None:
            return actual
        if isinstance(desired, str):
            desired_list = [desired]
        elif hasattr(desired, "tolist"):
            desired_list = list(desired.tolist())
        elif isinstance(desired, (list, tuple, set)):
            desired_list = list(desired)
        else:
            desired_list = [str(desired)]
        if len(desired_list) == 0:
            return actual
        matched = []
        desired_lower = [str(s).strip().lower() for s in desired_list]
        actual_lower_map = {s.strip().lower(): s for s in actual}
        for d_lower in desired_lower:
            if d_lower in actual_lower_map:
                matched.append(actual_lower_map[d_lower])
        if not matched:
            for d_lower in desired_lower:
                for actual_sheet in actual:
                    if d_lower in actual_sheet.strip().lower() or actual_sheet.strip().lower() in d_lower:
                        if actual_sheet not in matched:
                            matched.append(actual_sheet)
        if not matched:
            matched = actual
        return matched
    except Exception as e:
        ui_log(f"Error resolving sheets: {str(e)}", "error")
        return []

def _read_excel_raw(path_or_file, sheet_names: List[str], header) -> Dict[str, pd.DataFrame]:
    try:
        data = pd.read_excel(path_or_file, sheet_name=sheet_names, engine="openpyxl", header=header)
        if isinstance(data, pd.DataFrame):
            data = {sheet_names[0]: data}
        return data
    except Exception as e:
        ui_log(f"Error reading Excel sheets: {str(e)}", "error")
        return {}

# ----------------------------
# NASCAR CLEANER (FALLBACK)
# ----------------------------
def clean_columns_nascar(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    try:
        df = pd.DataFrame(df)

        def score_header(row_vals: List[str]) -> int:
            try:
                vals = [str(x).strip() for x in row_vals if pd.notna(x)]
                if not vals:
                    return -1
                s = 0
                low = [v.lower() for v in vals]
                if any(v == "driver" for v in low): s += 5
                if any(("dk" in v) and ("sal" in v or "pro" in v or "proj" in v) for v in low): s += 3
                if any(("fd" in v) and ("sal" in v or "pro" in v or "proj" in v) for v in low): s += 3
                if any(("qual" in v) or ("start" in v) for v in low): s += 2
                s -= sum(v.startswith("unnamed") for v in low)
                s += sum(v and not v.startswith("unnamed") for v in low)//2
                return s
            except:
                return -1

        best_idx, best_score = None, -10
        max_scan = min(8, len(df))
        for i in range(max_scan):
            try:
                row_data = df.iloc[i].tolist()
                sc = score_header(row_data)
                if sc > best_score:
                    best_idx, best_score = i, sc
            except:
                continue

        if best_idx is None or best_score < 1:
            best_idx, best_fill = 0, -1
            for i in range(max_scan):
                try:
                    vals = [str(x).strip() for x in df.iloc[i].tolist() if pd.notna(x)]
                    fill = sum(v and not v.lower().startswith("unnamed") for v in vals)
                    if fill > best_fill:
                        best_idx, best_fill = i, fill
                except:
                    continue

        try:
            header_vals = [str(x).strip() for x in df.iloc[best_idx].tolist()]
            df = df.iloc[best_idx + 1:].copy()
            df.columns = header_vals
        except Exception:
            return pd.DataFrame()

        valid_columns = [c for c in df.columns if c and not str(c).lower().startswith("unnamed") and str(c).strip()]
        if not valid_columns:
            return pd.DataFrame()
        df = df.loc[:, valid_columns]

        renames = {
            "Qual": "Start Pos",
            "qual": "Start Pos",
            "Start": "Start Pos",
            "DK Pro": "DK Proj",
            "FD Pro": "FD Proj",
            "DK Proj.": "DK Proj",
            "FD Proj.": "FD Proj",
            "Proj Finish": "Proj Fin",
            "ProjFin": "Proj Fin",
            "LL": "pLL",
            "DK Salary": "DK Sal",
            "FD Salary": "FD Sal",
        }
        df = df.rename(columns={c: renames.get(c, c) for c in df.columns})

        if "Driver" not in df.columns:
            for cand in ["Player Name", "Player", "Name"]:
                if cand in df.columns:
                    df = df.rename(columns={cand: "Driver"})
                    break

        keep = [c for c in df.columns if not any(rx.search(str(c)) for rx in EXCLUDE_PATTERNS)]
        df = df[keep].copy()
        df.columns = [str(c).strip() for c in df.columns]

        new_columns, seen = [], {}
        for col in df.columns:
            if col in seen:
                seen[col] += 1
                new_columns.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                new_columns.append(col)
        df.columns = new_columns

        for col in df.columns:
            try:
                cl = col.lower()
                if col in NUMERIC_KEYS or any(k in cl for k in ["proj", "sal", "own", "val", "finish", "laps"]):
                    series_data = df[col].astype(str).str.replace(r"[%$,]", "", regex=True)
                    series_data = series_data.replace(['', 'nan', 'None'], np.nan)
                    df[col] = pd.to_numeric(series_data, errors="coerce")
            except Exception:
                continue

        return df
    except Exception as e:
        ui_log(f"Error cleaning NASCAR columns: {str(e)}", "error")
        return pd.DataFrame()
    
def _mlb_find_header_row(df: pd.DataFrame) -> int:
    """
    Scan the first ~8 rows and pick the row that looks like the real MLB header.
    Targets short labels like V, BO, Player, DK Sal, FD Sal, Team, Opp, DK Proj, FD Proj, DK F, DK C, etc.
    Returns 0-based row index to use as header.
    """
    if df is None or df.empty:
        return 0

    target_tokens = {
        "v","bo","player","pos","dk sal","fd sal","team","opp",
        "dk proj","fd proj","dk val","fd val","dk pown","fd pown",
        "dk f","dk c","fd f","fd c","dk rtg","fd rtg",
        # pitchers/batters
        "ip","er","k","bb","hr","w","ab","1b","2b","3b","rbi","r","sb","h"
    }

    def score(row_vals):
        vals = [str(x).strip().lower() for x in row_vals if str(x).strip()]
        if not vals: 
            return -10
        s = 0
        for v in vals:
            if v in target_tokens: s += 3
            if "dk" in v or "fd" in v: s += 1
            if v.startswith("unnamed"): s -= 2
        return s

    best_idx, best_score = 0, -10
    scan = min(8, len(df))
    for i in range(scan):
        sc = score(df.iloc[i].tolist())
        if sc > best_score:
            best_idx, best_score = i, sc
    return best_idx


# ----------------------------
# DATA LOADING (CACHED)
# ----------------------------
@st.cache_data(show_spinner=False, ttl=60)
def load_data_for_sport(sport: str, path_or_file, only_sheets: Optional[List[str]] = None) -> Dict[str, pd.DataFrame]:
    try:
        name = getattr(path_or_file, "name", str(path_or_file))
        ext = os.path.splitext(name)[1].lower()

        if ext == ".csv" or ext == "":
            df = pd.read_csv(path_or_file)
            return {"Data": clean_columns(df)}

        sheet_names = resolve_allowed_sheets(path_or_file, only_sheets)
        if not sheet_names:
            ui_log(f"No sheets found in {name}", "warning")
            return {}

        ui_log(f"Found sheets: {sheet_names}", "info")
        out: Dict[str, pd.DataFrame] = {}

        if sport == "NASCAR":
            for sheet in sheet_names:
                ui_log(f"Processing sheet: {sheet}", "info")
                s = sheet.strip().lower()
                fast_df = None
                try:
                    if "betting" in s or "dashboard" in s:
                        fast_df = _fast_read_nascar_sheet(path_or_file, sheet, DESIRED_NASCAR_BETTING)
                    elif "proj" in s:
                        fast_df = _fast_read_nascar_sheet(path_or_file, sheet, DESIRED_NASCAR_PROJECTIONS)
                except Exception as e:
                    ui_log(f"Fast read failed for {sheet}: {str(e)}", "warning")

                if fast_df is not None and not fast_df.empty:
                    out[sheet] = fast_df
                    ui_log(f"Successfully loaded {sheet} using fast reader", "success")
                    continue

                try:
                    raw = _read_excel_raw(path_or_file, [sheet], header=None)
                    if not raw or sheet not in raw:
                        ui_log(f"Could not read raw data for sheet: {sheet}", "warning")
                        continue
                    cleaned = clean_columns_nascar(raw[sheet], sheet)

                    if "betting" in s or "dashboard" in s:
                        if "Driver" not in cleaned.columns:
                            for cand in ["Player Name", "Player", "Name"]:
                                if cand in cleaned.columns:
                                    cleaned = cleaned.rename(columns={cand: "Driver"})
                                    break
                        cleaned = cleaned.rename(columns={"Proj Finish": "Proj Fin", "ProjFin": "Proj Fin"})
                        available_betting_cols = [col for col in DESIRED_NASCAR_BETTING if col in cleaned.columns]
                        cleaned = cleaned.reindex(columns=available_betting_cols)

                    elif "proj" in s:
                        if "Driver" not in cleaned.columns:
                            for cand in ["Player Name", "Player", "Name"]:
                                if cand in cleaned.columns:
                                    cleaned = cleaned.rename(columns={cand: "Driver"})
                                    break
                        cleaned = cleaned.rename(columns={"Proj Finish": "Proj Fin", "ProjFin": "Proj Fin"})
                        available_proj_cols = [col for col in DESIRED_NASCAR_PROJECTIONS if col in cleaned.columns]
                        cleaned = cleaned.reindex(columns=available_proj_cols)

                    if not cleaned.empty:
                        out[sheet] = cleaned
                        ui_log(f"Successfully loaded {sheet} using fallback method", "success")
                    else:
                        ui_log(f"Sheet {sheet} resulted in empty DataFrame", "warning")
                except Exception as e:
                    ui_log(f"Failed to process sheet {sheet}: {str(e)}", "error")
            return out

            # ----------------------------
            # MLB loader (NEW)
            # ----------------------------
            if sport == "MLB":
                # Your real headers are on Excel row 2 → header=1 (0-indexed)
                raw = _read_excel_raw(path_or_file, sheet_names, header=1)
                out = {}
                for sheet, raw_df in raw.items():
                    dfc = clean_columns(raw_df)
                    dfc = _apply_aliases(dfc, _MLB_ALIAS)

                    s = sheet.strip().lower()

                    if "pitch" in s:
                        desired = [c for c in DESIRED_MLB_PITCHERS if c in dfc.columns]
                        if desired:
                            dfc = dfc.reindex(columns=desired)

                    elif "batter" in s or "hit" in s:
                        desired = [c for c in DESIRED_MLB_BATTERS if c in dfc.columns]
                        if desired:
                            dfc = dfc.reindex(columns=desired)

                    elif "stack" in s:
                        desired = [c for c in DESIRED_MLB_STACKS if c in dfc.columns]
                        if desired:
                            dfc = dfc.reindex(columns=desired)

                    out[sheet] = dfc
                return out


        # Fallback: NFL/general
        raw = _read_excel_raw(path_or_file, sheet_names, header=0)
        for sheet, raw_df in raw.items():
            out[sheet] = clean_columns(raw_df)
        return out

    except Exception as e:
        ui_log(f"Error loading data: {str(e)}", "error")
        return {}

# ----------------------------
# FORMATTING HELPERS
# ----------------------------
def format_currency(x):
    if pd.isna(x):
        return ""
    try:
        return f"${int(float(x)):,}"
    except (ValueError, TypeError):
        return str(x)

def format_percentage(x):
    if pd.isna(x):
        return ""
    try:
        val = float(x)
        return f"{val*100:.1f}%" if abs(val) <= 1 else f"{val:.1f}%"
    except (ValueError, TypeError):
        return str(x)

def format_number(x):
    if pd.isna(x):
        return ""
    try:
        val = float(x)
        return f"{int(val)}" if val == int(val) else f"{val:.1f}"
    except (ValueError, TypeError):
        return str(x)

def get_formatters(df: pd.DataFrame):
    fmt = {}
    numeric_like = {"o/u", "spread", "imp. tot", "qb", "rb", "wr", "te", "dst", "total"}
    for col in df.columns:
        col_lower = col.lower()
        if col_lower.endswith(" sal"):
            fmt[col] = format_currency
            continue
        if ("%" in col) or any(k in col_lower for k in ["own", "pown", "ownership", "tgt share"]):
            fmt[col] = format_percentage
            continue
        if (
            col_lower in numeric_like
            or any(col_lower.endswith(suf) for suf in ("proj", "val", "rtg", "floor", "ceiling", "pp", "dom"))
            or pd.api.types.is_numeric_dtype(df[col])
        ):
            fmt[col] = format_number
    return fmt

def _round_numeric_for_display(df_in: pd.DataFrame) -> pd.DataFrame:
    if df_in is None or not isinstance(df_in, pd.DataFrame):
        return pd.DataFrame()
    if df_in.empty:
        return df_in.copy()

    df_out = df_in.copy()

    keep_int = {"Win", "T3", "T5", "T10"}
    percent_markers = ("%", "own", "pown", "ownership", "tgt share", "opt", "lev")
    TEXT_NAME_GUARDS = {
        "driver", "player name", "player", "name",
        "wr1", "wr2", "wr3", "qb", "rb", "te",
        "bringback", "bring back", "stack", "stack team",
        "team", "opp", "wr", "wr1 name", "wr2 name", "wr3 name"
    }

    for c in df_out.columns:
        cl = str(c).lower()

        # --- SALARIES: scrub → numeric float → no NaN/∞ → round 0 ---
        if cl.endswith(" sal") or "salary" in cl or "price" in cl:
            s = (
                pd.Series(df_out[c])
                .astype(str)
                .str.replace(r"[^\d\.\-]", "", regex=True)   # drop $, commas, spaces
            )
            s = pd.to_numeric(s, errors="coerce")            # to float
            s = s.replace([np.inf, -np.inf], np.nan)         # remove infinities
            s = s.fillna(0.0).astype("float64").round(0)     # clean float dtype
            df_out[c] = s
            continue

        # keep certain columns as ints (Win, T3, T5, T10)
        if c in keep_int:
            df_out[c] = pd.to_numeric(df_out[c], errors="coerce").fillna(0).round(0).astype(float)
            continue

        # PERCENTS: coerce → auto-scale to 0–100 if needed → round 1
        if any(marker in cl for marker in percent_markers):
            vals = pd.to_numeric(df_out[c], errors="coerce")
            if pd.notna(vals).any() and vals.max(skipna=True) <= 1:
                vals = vals * 100.0
            df_out[c] = vals.fillna(0).astype(float).round(1)
            continue

        # Skip obvious text columns
        if any(token in cl for token in TEXT_NAME_GUARDS):
            continue

        # Generic numeric rounding (keep float dtype)
        if pd.api.types.is_numeric_dtype(df_out[c]):
            df_out[c] = pd.to_numeric(df_out[c], errors="coerce").fillna(0).astype(float).round(1)

    return df_out
# ----------------------------
# Streamlit column config helper
# ----------------------------
# ----------------------------
# Streamlit column config helper
# ----------------------------
def build_column_config(df: pd.DataFrame) -> dict:
    import streamlit as st
    cfg = {}

    def is_money(col: str) -> bool:
        cl = str(col).lower()
        # treat any "* Sal" or "* Salary"/"Price" as salary
        return cl.endswith(" sal") or "salary" in cl or "price" in cl

    def is_percent(col: str) -> bool:
        cl = str(col).lower()
        return ("%" in col) or any(k in cl for k in ["own", "pown", "ownership", "opt", "lev", "tgt share"])

    for col in df.columns:
        try:
            if is_money(col):
                # plain integer (no $ or commas)
                cfg[col] = st.column_config.NumberColumn(
                    col,
                    format="%.0f",
                    step=100.0,   # tweak if you prefer different step size
                    min_value=0.0
                )
            elif is_percent(col) and pd.api.types.is_numeric_dtype(df[col]):
                cfg[col] = st.column_config.NumberColumn(col, format="%.1f%%")
            elif pd.api.types.is_numeric_dtype(df[col]):
                cfg[col] = st.column_config.NumberColumn(col, format="%.1f")
        except Exception:
            pass
    return cfg

def _format_for_display(df_in: pd.DataFrame) -> pd.DataFrame:
    if df_in is None or not isinstance(df_in, pd.DataFrame):
        return pd.DataFrame()
    if df_in.empty:
        return df_in.copy()

    df = df_in.copy()
    for col in df.columns:
        cl = str(col).lower()
        try:
            if cl.endswith(" sal") or cl.endswith(" salary") or ("salary" in cl) or ("price" in cl):
                df[col] = df[col].apply(format_currency)
            elif ("%" in col) or any(k in cl for k in ["own", "pown", "ownership", "tgt share", "opt%"]):
                df[col] = df[col].apply(format_percentage)
            elif any(tok in cl for tok in MLB_THREE_DEC_STATS):
                df[col] = pd.to_numeric(df[col], errors="coerce").apply(lambda v: "" if pd.isna(v) else f"{v:.3f}")
            elif pd.api.types.is_numeric_dtype(df[col]):
                df[col] = df[col].apply(format_number)
        except Exception:
            pass
    return df

# Tiny extra format helpers (used in Position Summary)
def _fmt_currency0(x):
    x = pd.to_numeric(x, errors="coerce")
    return "" if pd.isna(x) else f"{int(round(x))}"

def _fmt_number1(x):
    x = pd.to_numeric(x, errors="coerce")
    return "" if pd.isna(x) else f"{x:.1f}"

def _fmt_percent1(x):
    x = pd.to_numeric(x, errors="coerce")
    if pd.isna(x): return ""
    if abs(x) <= 1: x *= 100
    return f"{x:.1f}%"

# ======================================================
# CONTEXT + COLUMN HELPERS & CHART BUILDERS
# ======================================================
def _lc_set(cols): return {str(c).strip().lower() for c in cols}

def is_nascar_context(sport: str, sheet_name: str, df: pd.DataFrame) -> bool:
    if sport.upper() == "NASCAR":
        return True
    hints = {"driver", "proj fin", "qual", "win%", "t3%", "t5%", "t10%", "dk dom", "fd dom"}
    return any(h in _lc_set(df.columns) for h in hints)

def is_stacks_context(sheet_name: str, df: pd.DataFrame) -> bool:
    s = str(sheet_name).strip().lower()
    if "stack" in s:
        return True
    role_cols = {"qb", "wr", "wr1", "wr2", "te", "rb", "bringback", "bring back", "stack", "stack team", "team"}
    return len(_lc_set(df.columns).intersection(role_cols)) >= 2

def is_nfl_projections_context(sport: str, sheet_name: str, df: pd.DataFrame) -> bool:
    return (sport.upper() == "NFL") and (not is_stacks_context(sheet_name, df))

# MLB context (NEW)
def is_mlb_context(sport: str, sheet_name: str, df: pd.DataFrame) -> bool:
    if sport.upper() == "MLB":
        return True
    hints = {"bat order","bats","pitcher hand","k proj","ip proj","xfip","woba","ops","team imp. tot"}
    cols = _lc_set(df.columns)
    return any(h in cols for h in hints)

def coalesce(df: pd.DataFrame, *cands: str) -> Optional[str]:
    cols = list(df.columns)
    lowers = {str(c).strip().lower(): c for c in cols}
    for c in cands:
        lc = c.strip().lower()
        if lc in lowers:
            return lowers[lc]
        for col in cols:
            if str(col).strip().lower() == lc:
                return col
    return None

def exists_all(df: pd.DataFrame, *cols: str) -> bool:
    present = _lc_set(df.columns)
    return all(c.strip().lower() in present for c in cols)

# --- Trend helper ---
def _add_linear_trend(fig: go.Figure, x: pd.Series, y: pd.Series, name: str = "Trend") -> None:
    try:
        xv = pd.to_numeric(x, errors="coerce")
        yv = pd.to_numeric(y, errors="coerce")
        use = pd.DataFrame({"x": xv, "y": yv}).dropna()
        if len(use) < 3:
            return
        z = np.polyfit(use["x"].values, use["y"].values, 1)
        p = np.poly1d(z)
        xs = np.linspace(use["x"].min(), use["x"].max(), 50)
        fig.add_trace(go.Scatter(x=xs, y=p(xs), mode="lines", name=name, line=dict(dash="dot")))
    except Exception:
        pass

# ----------------------------
# NASCAR CHARTS
# ----------------------------
def nascar_salary_vs_proj(df: pd.DataFrame, site: str) -> go.Figure:
    sal = coalesce(df, f"{site} Sal")
    proj = coalesce(df, f"{site} Proj")
    if not sal or not proj:
        return go.Figure()
    use = df[[sal, proj] + [c for c in ["Driver"] if c in df.columns]].dropna()
    if use.empty:
        return go.Figure()
    fig = px.scatter(
        use, x=sal, y=proj, hover_name="Driver" if "Driver" in use.columns else None,
        title=f"NASCAR — {site} Salary vs {site} Projection",
        labels={sal:"Salary ($)", proj:"Projection"}
    )
    _add_linear_trend(fig, use[sal], use[proj])
    fig.update_layout(height=460)
    return fig

def nascar_qual_vs_proj(df: pd.DataFrame, site: str) -> go.Figure:
    qual = coalesce(df, "Qual")
    proj = coalesce(df, f"{site} Proj")
    if not qual or not proj:
        return go.Figure()
    use = df[[qual, proj] + [c for c in ["Driver"] if c in df.columns]].dropna()
    if use.empty:
        return go.Figure()
    fig = px.scatter(
        use, x=qual, y=proj, hover_name="Driver" if "Driver" in use.columns else None,
        title=f"NASCAR — Qualifying Position vs {site} Projection",
        labels={qual:"Qualifying (Start)", proj:"Projection"}
    )
    _add_linear_trend(fig, use[qual], use[proj])
    fig.update_layout(height=460)
    return fig

def nascar_opt_vs_own(df: pd.DataFrame, site: str) -> go.Figure:
    opt = coalesce(df, f"{site} Opt%")
    own = coalesce(df, f"{site} pOWN%")
    if not opt or not own:
        return go.Figure()
    keep = [opt, own] + [c for c in ["Driver", f"{site} Proj"] if c in df.columns]
    use = df[keep].dropna()
    if use.empty:
        return go.Figure()
    hdata = {own:":.1f", opt:":.1f"}
    if f"{site} Proj" in use.columns:
        hdata[f"{site} Proj"] = ":.1f"
    fig = px.scatter(
        use, x=own, y=opt, hover_name="Driver" if "Driver" in use.columns else None,
        hover_data=hdata,
        title=f"NASCAR — {site} Optimal% vs {site} pOWN%",
        labels={own:f"{site} pOWN%", opt:f"{site} Optimal%"},
    )
    fig.update_layout(height=460)
    return fig

# ----------------------------
# NFL PLAYER-LEVEL CHARTS
# ----------------------------
def nfl_salary_vs_proj(df: pd.DataFrame, site: str) -> go.Figure:
    sal, proj = coalesce(df, f"{site} Sal"), coalesce(df, f"{site} Proj")
    if not sal or not proj: return go.Figure()
    extras = [c for c in ["Pos", "Player Name", "Team", "Opp"] if c in df.columns]
    use = df[[sal, proj] + extras].dropna()
    if use.empty: return go.Figure()
    fig = px.scatter(
        use, x=sal, y=proj, color="Pos" if "Pos" in use.columns else None,
        hover_name="Player Name" if "Player Name" in use.columns else None,
        title=f"NFL — {site} Salary vs Projection",
        labels={sal:"Salary ($)", proj:"Projection"},
        color_discrete_map=POSITION_COLORS if "Pos" in use.columns else None
    )
    _add_linear_trend(fig, use[sal], use[proj])
    fig.update_layout(height=460, showlegend="Pos" in use.columns)
    return fig

def nfl_proj_vs_own(df: pd.DataFrame, site: str) -> go.Figure:
    proj, own = coalesce(df, f"{site} Proj"), coalesce(df, f"{site} pOWN%")
    val = coalesce(df, f"{site} Val")
    if not proj or not own: return go.Figure()
    keep = [proj, own] + [c for c in ["Pos", "Player Name", "Team", val] if c]
    use = df[keep].dropna()
    if use.empty: return go.Figure()
    size = (use[val] if val else pd.Series([8]*len(use))).abs().clip(1, None)
    fig = px.scatter(
        use, x=proj, y=own, size=size,
        color="Pos" if "Pos" in use.columns else None,
        hover_name="Player Name" if "Player Name" in use.columns else None,
        title=f"NFL — {site} Projection vs {site} pOWN%",
        labels={proj:"Projection", own:"pOWN%"},
        color_discrete_map=POSITION_COLORS if "Pos" in use.columns else None
    )
    fig.update_traces(marker=dict(line=dict(width=0)))
    _add_linear_trend(fig, use[proj], use[own])
    fig.update_layout(height=460, showlegend="Pos" in use.columns)
    return fig

def nfl_val_vs_proj(df: pd.DataFrame, site: str) -> go.Figure:
    val, proj = coalesce(df, f"{site} Val"), coalesce(df, f"{site} Proj")
    if not val or not proj: return go.Figure()
    use = df[[val, proj] + [c for c in ["Pos", "Player Name"] if c in df.columns]].dropna()
    if use.empty: return go.Figure()
    fig = px.scatter(
        use, x=val, y=proj,
        color="Pos" if "Pos" in use.columns else None,
        hover_name="Player Name" if "Player Name" in use.columns else None,
        title=f"NFL — {site} Value vs {site} Projection",
        labels={val:"Value", proj:"Projection"},
        color_discrete_map=POSITION_COLORS if "Pos" in use.columns else None
    )
    _add_linear_trend(fig, use[val], use[proj])
    fig.update_layout(height=460, showlegend="Pos" in use.columns)
    return fig

def nfl_pos_box(df: pd.DataFrame, site: str, metric: str = "Proj") -> go.Figure:
    mcol = coalesce(df, f"{site} {metric}")
    if "Pos" not in df.columns or not mcol: return go.Figure()
    use = df[["Pos", mcol]].dropna()
    if use.empty: return go.Figure()
    fig = px.box(use, x="Pos", y=mcol, points="suspectedoutliers",
                 title=f"NFL — {site} {metric} by Position", labels={mcol:metric})
    fig.update_layout(height=460)
    return fig

# ----------------------------
# MLB CHARTS (NEW)
# ----------------------------
def mlb_bat_order_vs_proj(df: pd.DataFrame, site: str) -> go.Figure:
    order = coalesce(df, "Bat Order")
    proj  = coalesce(df, f"{site} Proj")
    if not order or not proj: return go.Figure()
    use = df[[order, proj] + [c for c in ["Player Name", "Team", "Pos"] if c in df.columns]].dropna()
    if use.empty: return go.Figure()
    jitter = (np.random.rand(len(use)) - 0.5) * 0.08
    fig = px.scatter(use, x=pd.to_numeric(use[order], errors="coerce")+jitter, y=proj,
                     color="Pos" if "Pos" in use.columns else None,
                     hover_name="Player Name" if "Player Name" in use.columns else None,
                     title=f"MLB — Bat Order vs {site} Projection",
                     labels={order:"Bat Order", proj:"Projection"})
    fig.update_layout(height=420, showlegend="Pos" in use.columns)
    return fig

def mlb_teamimp_vs_proj(df: pd.DataFrame, site: str) -> go.Figure:
    imp  = coalesce(df, "Team Imp. Tot")
    proj = coalesce(df, f"{site} Proj")
    if not imp or not proj: return go.Figure()
    use = df[[imp, proj] + [c for c in ["Player Name","Team"] if c in df.columns]].dropna()
    if use.empty: return go.Figure()
    fig = px.scatter(use, x=imp, y=proj,
                     hover_name="Player Name" if "Player Name" in use.columns else None,
                     color="Team" if "Team" in use.columns else None,
                     title=f"MLB — Team Implied Total vs {site} Projection",
                     labels={imp:"Team Implied Total", proj:"Projection"})
    _add_linear_trend(fig, use[imp], use[proj])
    fig.update_layout(height=420, showlegend="Team" in use.columns)
    return fig

def mlb_salary_vs_kproj(df: pd.DataFrame, site: str) -> go.Figure:
    sal = coalesce(df, f"{site} Sal")
    kp  = coalesce(df, "K Proj")
    if not sal or not kp: return go.Figure()
    use = df[[sal, kp] + [c for c in ["Player Name","Team"] if c in df.columns]].dropna()
    if use.empty: return go.Figure()
    fig = px.scatter(use, x=sal, y=kp,
                     hover_name="Player Name" if "Player Name" in use.columns else None,
                     color="Team" if "Team" in use.columns else None,
                     title=f"MLB — {site} Salary vs K Proj",
                     labels={sal:"Salary ($)", kp:"K Proj"})
    _add_linear_trend(fig, use[sal], use[kp])
    fig.update_layout(height=420, showlegend="Team" in use.columns)
    return fig

# ----------------------------
# NFL STACKS CHARTS (with TOTAL)
# ----------------------------
def stacks_find_cols(df: pd.DataFrame, site: str):
    team = coalesce(df, "Team", "team", "Stack Team", "StackTeam")

    # Total (overall stack projection total)
    total = coalesce(
        df, "Total", "Stack Total", "Team Total", "Total Proj",
        "Total Projection", "Sum Proj", "Sum Projection", "Total Proj"
    )

    # Implied total (Vegas)
    imp_tot = coalesce(
        df, "Imp. Tot", "Imp Tot", "Implied Total", "Team Implied",
        "Vegas Team Total", "Vegas Total", "Team Imp. Tot"
    )

    # Projection (fallbacks)
    proj = coalesce(
        df,
        "Stack Proj", "Stack Projection", "Team Stack Proj", "Team Stack Projection",
        f"{site} Stack Proj",
        "Total Stack Proj", "Total Stack Projection",
        f"{site} Proj", "Proj", "Projection",
        total
    )

    own = coalesce(
        df,
        f"{site} Stack pOWN%", f"{site} pOWN%", "Stack pOWN%", "pOWN%",
        "Ownership%", "Own%", "Ownership"
    )
    opt = coalesce(
        df,
        f"{site} Stack Opt%", f"{site} Opt%", "Stack Opt%", "Opt%", "Optimal%", "Optimal %"
    )
    salary = coalesce(
        df,
        "Stack Salary", f"{site} Stack Salary", f"{site} Sal", f"{site} Stack Sal",
        "Salary", "Stack Price", "Total Salary", "Price", "Total Price",
        f"{site} Stack Salary"
    )
    return team, proj, own, opt, salary, total, imp_tot

def stacks_total_hist(df: pd.DataFrame) -> go.Figure:
    total = coalesce(df, "Total", "Stack Total", "Total Proj", "Total Projection", "Sum Proj", "Total Proj")
    if not total: return go.Figure()
    use = df[[total]].dropna()
    if use.empty: return go.Figure()
    fig = px.histogram(use, x=total, nbins=30, title="Stacks — Total Projection Distribution", labels={total:"Total Projection"})
    fig.update_layout(height=420)
    return fig

def stacks_total_vs_salary(df: pd.DataFrame, site: str) -> go.Figure:
    _, _, _, _, salary, total, _ = stacks_find_cols(df, site)
    if not total or not salary: return go.Figure()
    keep = [total, salary] + [c for c in ["Team"] if c in df.columns]
    use = df[keep].dropna()
    if use.empty: return go.Figure()
    fig = px.scatter(
        use, x=salary, y=total, color="Team" if "Team" in use.columns else None,
        title=f"Stacks — {site} Stack Salary vs TOTAL Projection",
        labels={salary:"Stack Salary ($)", total:"Total Projection"},
    )
    _add_linear_trend(fig, use[salary], use[total])
    fig.update_layout(height=460, showlegend="Team" in use.columns)
    return fig

def stacks_total_vs_imptot(df: pd.DataFrame) -> go.Figure:
    total = coalesce(df, "Total", "Stack Total", "Total Proj", "Total Projection", "Sum Proj", "Total Proj")
    imp   = coalesce(df, "Imp. Tot", "Imp Tot", "Implied Total", "Team Implied", "Vegas Team Total", "Vegas Total", "Team Imp. Tot")
    if not total or not imp: return go.Figure()
    use = df[[total, imp] + [c for c in ["Team"] if c in df.columns]].dropna()
    if use.empty: return go.Figure()
    fig = px.scatter(
        use, x=imp, y=total, color="Team" if "Team" in use.columns else None,
        title="Stacks — TOTAL Projection vs Implied Total",
        labels={imp:"Implied Total", total:"Total Projection"},
    )
    _add_linear_trend(fig, use[imp], use[total])
    fig.update_layout(height=460, showlegend="Team" in use.columns)
    return fig

def stacks_total_vs_own(df: pd.DataFrame, site: str) -> go.Figure:
    total = coalesce(df, "Total", "Stack Total", "Total Proj", "Total Projection", "Sum Proj", "Total Proj")
    own   = coalesce(df, f"{site} Stack pOWN%", f"{site} pOWN%", "Stack pOWN%", "pOWN%", "Own%")
    if not total or not own: return go.Figure()
    keep = [total, own] + [c for c in ["Team"] if c in df.columns]
    use = df[keep].dropna()
    if use.empty: return go.Figure()
    fig = px.scatter(
        use, x=own, y=total, color="Team" if "Team" in use.columns else None,
        title=f"Stacks — TOTAL Projection vs {site} pOWN%",
        labels={own:f"{site} pOWN%", total:"Total Projection"},
    )
    fig.update_layout(height=460, showlegend="Team" in use.columns)
    return fig

def stacks_opt_vs_own(df: pd.DataFrame, site: str) -> go.Figure:
    team, proj, own, opt, salary, _, _ = stacks_find_cols(df, site)
    if not own or not opt: return go.Figure()
    keep = [own, opt] + [c for c in [proj, team, salary] if c]
    use = df[keep].dropna()
    if use.empty: return go.Figure()
    hdata = {own:":.1f", opt:":.1f"}
    if proj:   hdata[proj] = ":.1f"
    if salary: hdata[salary] = ":,.0f"
    fig = px.scatter(
        use, x=own, y=opt, color=team if team else None,
        hover_data=hdata,
        title=f"Stacks — {site} Stack Optimal% vs {site} Stack pOWN%",
        labels={own:f"{site} pOWN%", opt:f"{site} Optimal%"},
    )
    fig.update_layout(height=460, showlegend=bool(team))
    return fig

def stacks_table(df: pd.DataFrame, site: str) -> pd.DataFrame:
    team, proj, own, opt, salary, total, _ = stacks_find_cols(df, site)
    cols = [c for c in [team, total or proj, own, opt, salary] if c]
    if not cols: return pd.DataFrame()
    out = df[cols].copy()
    ren = {}
    if team: ren[team] = "Team"
    if total or proj: ren[total or proj] = "Total Proj" if total else "Proj"
    if own:  ren[own]  = f"{site} pOWN%"
    if opt:  ren[opt]  = f"{site} Opt%"
    if salary: ren[salary] = "Stack Salary"
    out = out.rename(columns=ren)
    if f"{site} Opt%" in out.columns and f"{site} pOWN%" in out.columns:
        out[f"{site} Lev%"] = out[f"{site} Opt%"] - out[f"{site} pOWN%"]
    sort_col = "Total Proj" if "Total Proj" in out.columns else ("Proj" if "Proj" in out.columns else None)
    if sort_col:
        out = out.sort_values(sort_col, ascending=False)
    return out

# ----------------------------
# ANALYTICS RENDERER (UNIFORM DK/FD) — with MLB
# ----------------------------
def render_analytics_auto(df: pd.DataFrame, selected_sport: str, selected_sheet: str, site_filter: str):
    sites = ["DK", "FD"] if site_filter == "Both" else [site_filter]

    # MLB
    if is_mlb_context(selected_sport, selected_sheet, df):
        st.subheader("📈 MLB Analytics")
        for site in sites:
            st.markdown(f"**{site}**")
            c1, c2 = st.columns(2)
            with c1:
                fig = nfl_salary_vs_proj(df, site)   # shared pattern
                if fig.data: st.plotly_chart(fig, use_container_width=True)
            with c2:
                fig = nfl_proj_vs_own(df, site)      # shared pattern
                if fig.data: st.plotly_chart(fig, use_container_width=True)

            c3, c4 = st.columns(2)
            with c3:
                fig = nfl_val_vs_proj(df, site)      # shared pattern
                if fig.data: st.plotly_chart(fig, use_container_width=True)
            with c4:
                sname = selected_sheet.strip().lower()
                if "pitch" in sname:
                    fig = mlb_salary_vs_kproj(df, site)
                else:
                    fig = mlb_bat_order_vs_proj(df, site)
                    if not fig.data:
                        fig = mlb_teamimp_vs_proj(df, site)
                if fig.data: st.plotly_chart(fig, use_container_width=True)
        return

    # NASCAR
    if is_nascar_context(selected_sport, selected_sheet, df):
        st.subheader("📈 NASCAR Analytics")
        for site in sites:
            c1, c2 = st.columns(2)
            with c1:
                fig = nascar_salary_vs_proj(df, site)
                if fig.data: st.plotly_chart(fig, use_container_width=True)
            with c2:
                fig = nascar_qual_vs_proj(df, site)
                if fig.data: st.plotly_chart(fig, use_container_width=True)
            c3, _ = st.columns(2)
            with c3:
                fig = nascar_opt_vs_own(df, site)
                if fig.data: st.plotly_chart(fig, use_container_width=True)
        return

    # NFL Stacks (general stacks logic)
    if is_stacks_context(selected_sheet, df):
        st.subheader("📈 Stacks Analytics")
        # Global (site-agnostic)
        a1, a2 = st.columns(2)
        with a1:
            fig = stacks_total_hist(df)
            if fig.data: st.plotly_chart(fig, use_container_width=True)
        with a2:
            fig = stacks_total_vs_imptot(df)
            if fig.data: st.plotly_chart(fig, use_container_width=True)
        # Per-site
        for site in sites:
            st.markdown(f"**{site}**")
            b1, b2 = st.columns(2)
            with b1:
                fig = stacks_total_vs_salary(df, site)
                if fig.data: st.plotly_chart(fig, use_container_width=True)
            with b2:
                fig = stacks_opt_vs_own(df, site)
                if fig.data: st.plotly_chart(fig, use_container_width=True)
        return

    # NFL Player Analytics (fallback)
    st.subheader("📈 NFL Player Analytics")
    for site in sites:
        st.markdown(f"**{site}**")
        c1, c2 = st.columns(2)
        with c1:
            fig = nfl_salary_vs_proj(df, site)
            if fig.data: st.plotly_chart(fig, use_container_width=True)
        with c2:
            fig = nfl_proj_vs_own(df, site)
            if fig.data: st.plotly_chart(fig, use_container_width=True)

        c3, c4 = st.columns(2)
        with c3:
            fig = nfl_val_vs_proj(df, site)
            if fig.data: st.plotly_chart(fig, use_container_width=True)
        with c4:
            fig = nfl_pos_box(df, site, metric="Proj")
            if fig.data: st.plotly_chart(fig, use_container_width=True)

# ----------------------------
# NFL-ONLY POSITION SUMMARY (unchanged)
# ----------------------------
def nfl_position_summary(df: pd.DataFrame, site: str) -> pd.DataFrame:
    if "Pos" not in df.columns:
        return pd.DataFrame()

    base_cols = {
        "Sal": f"{site} Sal",
        "Proj": f"{site} Proj",
        "Val": f"{site} Val",
        "Own": f"{site} pOWN%",
    }
    have = {k: v for k, v in base_cols.items() if v in df.columns}
    if not have:
        return pd.DataFrame()

    d = df[["Pos"] + list(have.values())].copy()

    for col in have.values():
        d[col] = pd.Series(d[col]).astype(str).str.replace(r"[%$,]", "", regex=True)
        d[col] = pd.to_numeric(d[col], errors="coerce")

    g = d.groupby("Pos").agg({v: ["count", "mean", "std"] for v in have.values()})
    g.columns = [f"{orig}_{stat}" for orig, stat in g.columns]
    out = g.reset_index()

    cols_display = ["Pos"]
    rows = []

    for _, row in out.iterrows():
        r = {"Pos": row["Pos"]}

        if base_cols.get("Sal") and f"{base_cols['Sal']}_count" in out.columns:
            r["Count"] = int(row[f"{base_cols['Sal']}_count"])
            mean_sal = row.get(f"{base_cols['Sal']}_mean", pd.NA)
            std_sal  = row.get(f"{base_cols['Sal']}_std", pd.NA)
            r[f"{site} Sal (mean)"] = _fmt_currency0(mean_sal)
            r[f"{site} Sal (std)"]  = _fmt_currency0(std_sal)
            if "Count" not in cols_display:
                cols_display.extend(["Count", f"{site} Sal (mean)", f"{site} Sal (std)"])

        if base_cols.get("Proj") and f"{base_cols['Proj']}_mean" in out.columns:
            r[f"{site} Proj (mean)"] = _fmt_number1(row[f"{base_cols['Proj']}_mean"])
            r[f"{site} Proj (std)"]  = _fmt_number1(row[f"{base_cols['Proj']}_std"])
            if f"{site} Proj (mean)" not in cols_display:
                cols_display.extend([f"{site} Proj (mean)", f"{site} Proj (std)"])

        if base_cols.get("Val") and f"{base_cols['Val']}_mean" in out.columns:
            r[f"{site} Val (mean)"] = _fmt_number1(row[f"{base_cols['Val']}_mean"])
            r[f"{site} Val (std)"]  = _fmt_number1(row[f"{base_cols['Val']}_std"])
            if f"{site} Val (mean)" not in cols_display:
                cols_display.extend([f"{site} Val (mean)", f"{site} Val (std)"])

        if base_cols.get("Own") and f"{base_cols['Own']}_mean" in out.columns:
            r[f"{site} pOWN% (mean)"] = _fmt_percent1(row[f"{base_cols['Own']}_mean"])
            r[f"{site} pOWN% (std)"]  = _fmt_percent1(row[f"{base_cols['Own']}_std"])
            if f"{site} pOWN% (mean)" not in cols_display:
                cols_display.extend([f"{site} pOWN% (mean)", f"{site} pOWN% (std)"])

        rows.append(r)

    disp = pd.DataFrame(rows, columns=cols_display)

    sort_key = f"{site} Proj (mean)"
    if sort_key in disp.columns:
        tmp = pd.to_numeric(disp[sort_key].str.replace(r"[^\d.]", "", regex=True), errors="coerce")
        disp = disp.iloc[tmp.sort_values(ascending=False).index].reset_index(drop=True)

    return disp

# ======================================================
# UI
# ======================================================

# Sidebar sport picker
st.sidebar.header("🎛️ Controls")
st.sidebar.markdown("---")
st.sidebar.checkbox("Show load logs", value=False, key="show_logs")

SPORTS = list(DEFAULT_SPORTS.keys())
selected_sport = st.sidebar.selectbox("🏅 Sport", SPORTS, index=0)

# Data status (tidy)
with st.expander("📊 Data Status", expanded=False):
    for item in DEFAULT_SPORTS.get(selected_sport, []):
        path = item["path"]; sheets = item.get("sheets", [])
        exists = os.path.exists(path)
        status_icon = "✅" if exists else "❌"
        st.markdown(f"{status_icon} **{item['label']}**")
        st.markdown(f"   📁 Path: {path}")
        st.markdown(f"   📋 Sheets (expected): {sheets}")

# State init
if "datasets" not in st.session_state:
    st.session_state.datasets = {}
if "visible_cols" not in st.session_state:
    st.session_state.visible_cols = {}
if "presets" not in st.session_state:
    st.session_state.presets = _load_preset_store()

# Auto-load local files
if selected_sport not in st.session_state.datasets:
    st.session_state.datasets[selected_sport] = {}

for item in DEFAULT_SPORTS.get(selected_sport, []):
    label, path = item["label"], item["path"]
    desired_sheets = item.get("sheets", [])
    if os.path.exists(path) and label not in st.session_state.datasets[selected_sport]:
        with st.spinner(f"Loading {selected_sport} — {label}..."):
            try:
                data = load_data_for_sport(selected_sport, path, only_sheets=desired_sheets)
                if data:
                    st.session_state.datasets[selected_sport][label] = {
                        "data": data,
                        "allowed": set(data.keys()) if (desired_sheets is not None and len(list(desired_sheets)) > 0) else None,
                    }
                    ui_log(f"Loaded {label} with sheets: {list(data.keys())}", "success")
                else:
                    ui_log(f"{label}: No matching sheets found or failed to load", "warning")
            except Exception as e:
                ui_log(f"Failed to load {label}: {str(e)}", "error")

# Uploads
with st.expander("📤 Upload Additional Files"):
    uploaded_files = st.file_uploader(
        f"Upload {selected_sport} CSV/Excel files",
        type=["csv", "xlsx", "xlsm", "xls"],
        accept_multiple_files=True,
        help="Upload additional data files to analyze",
    )
    if uploaded_files:
        for file in uploaded_files:
            base_name = os.path.splitext(file.name)[0]
            if base_name not in st.session_state.datasets[selected_sport]:
                try:
                    data = load_data_for_sport(selected_sport, file, only_sheets=None)
                    if data:
                        st.session_state.datasets[selected_sport][base_name] = {"data": data, "allowed": None}
                        ui_log(f"Uploaded {file.name}", "success")
                    else:
                        ui_log(f"Failed to process {file.name}", "warning")
                except Exception as e:
                    ui_log(f"Error uploading {file.name}: {str(e)}", "error")

# Guard
if not st.session_state.datasets.get(selected_sport):
    st.warning(f"⚠️ No {selected_sport} datasets available. Check file paths or upload files.")
    st.stop()

# Sidebar dataset/sheet pickers
dataset_options = list(st.session_state.datasets[selected_sport].keys())
selected_dataset = st.sidebar.selectbox("📊 Select Dataset", dataset_options)
dataset_entry = st.session_state.datasets[selected_sport][selected_dataset]
sheet_options = list(dataset_entry["data"].keys())

# Force MLB sheet order
if selected_sport == "MLB":
    preferred = ["Pitcher Projections", "Batter Projections", "Top Stacks"]
    preferred_present = [s for s in preferred if s in sheet_options]
    extras = sorted([s for s in sheet_options if s not in preferred])
    sheet_options = preferred_present + extras

selected_sheet = st.sidebar.selectbox("📋 Select Sheet", sheet_options)

df = dataset_entry["data"].get(selected_sheet)
if df is None or (isinstance(df, pd.DataFrame) and df.empty):
    st.warning("⚠️ Selected sheet is empty.")
    st.stop()

# Sidebar: advanced filters (DK) + MLB extras
st.sidebar.markdown("---")
st.sidebar.subheader("🔍 Advanced Filters")
if "DK Sal" in df.columns and df["DK Sal"].notna().any():
    _min_sal = int(pd.to_numeric(df["DK Sal"], errors="coerce").min())
    _max_sal = int(pd.to_numeric(df["DK Sal"], errors="coerce").max())
    min_sal, max_sal = st.sidebar.slider("DK Salary Range", min_value=_min_sal, max_value=_max_sal, value=(_min_sal, _max_sal))
else:
    min_sal, max_sal = None, None

if "DK Proj" in df.columns and df["DK Proj"].notna().any():
    _min_proj = float(pd.to_numeric(df["DK Proj"], errors="coerce").min())
    _max_proj = float(pd.to_numeric(df["DK Proj"], errors="coerce").max())
    min_proj, max_proj = st.sidebar.slider("DK Projection Range", min_value=_min_proj, max_value=_max_proj, value=(_min_proj, _max_proj), step=0.5)
else:
    min_proj, max_proj = None, None

# MLB-specific filters (conditional)
bat_min = bat_max = None
selected_bats = "All"
selected_pitch_hand = "All"
k_min = k_max = None
ip_min = ip_max = None

if selected_sport == "MLB":
    if "Bat Order" in df.columns and df["Bat Order"].notna().any():
        bo_series = pd.to_numeric(df["Bat Order"], errors="coerce").dropna()
        if not bo_series.empty:
            _min_bo, _max_bo = int(bo_series.min()), int(bo_series.max())
            bat_min, bat_max = st.sidebar.slider("Bat Order", min_value=max(1, _min_bo), max_value=min(9, _max_bo), value=(max(1, _min_bo), min(9, _max_bo)))
    if "Bats" in df.columns:
        bats_options = ["All"] + sorted([b for b in df["Bats"].dropna().astype(str).unique()])
        selected_bats = st.sidebar.selectbox("Bats", bats_options)
    if "Pitcher Hand" in df.columns:
        p_hand_options = ["All"] + sorted([b for b in df["Pitcher Hand"].dropna().astype(str).unique()])
        selected_pitch_hand = st.sidebar.selectbox("Pitcher Hand (vs)", p_hand_options)
    if "K Proj" in df.columns and df["K Proj"].notna().any():
        k_series = pd.to_numeric(df["K Proj"], errors="coerce").dropna()
        if not k_series.empty:
            kmin, kmax = float(k_series.min()), float(k_series.max())
            k_min, k_max = st.sidebar.slider("K Proj Range", min_value=kmin, max_value=kmax, value=(kmin, kmax), step=0.5)
    if "IP Proj" in df.columns and df["IP Proj"].notna().any():
        ip_series = pd.to_numeric(df["IP Proj"], errors="coerce").dropna()
        if not ip_series.empty:
            ipmin, ipmax = float(ip_series.min()), float(ip_series.max())
            ip_min, ip_max = st.sidebar.slider("IP Proj Range", min_value=ipmin, max_value=ipmax, value=(ipmin, ipmax), step=0.1)

# Main tabs
tab1, tab2, tab3 = st.tabs(["📊 Data Explorer", "📈 Analytics", "📋 Position Summary"])

with tab1:
    col1, col2 = st.columns([3, 1])
    with col2:
        st.subheader("📈 Quick Stats")
        st.metric("Total Rows", len(df))
        if "DK Sal" in df.columns and df["DK Sal"].notna().any():
            st.metric("Avg DK Salary", f"${pd.to_numeric(df['DK Sal'], errors='coerce').mean():,.0f}")
        if "DK Proj" in df.columns and df["DK Proj"].notna().any():
            st.metric("Avg DK Projection", f"{pd.to_numeric(df['DK Proj'], errors='coerce').mean():.1f}")
        if "Pos" in df.columns:
            positions = df["Pos"].value_counts()
            st.write("**Positions:**")
            for pos, count in positions.head(5).items():
                st.write(f"• {pos}: {count}")

    with col1:
        st.subheader(f"📊 {selected_sport} — {selected_dataset} — {selected_sheet}")

        filter_col1, filter_col2, filter_col3, filter_col4 = st.columns(4)
        with filter_col1:
            search_query = st.text_input("🔍 Enter player/driver name...", placeholder="Search...")
        with filter_col2:
            pos_options = ["All"] + (sorted(df["Pos"].dropna().astype(str).unique()) if "Pos" in df.columns else [])
            selected_pos = st.selectbox("🏃 Position", pos_options)
        with filter_col3:
            team_options = ["All"] + (sorted(df["Team"].dropna().astype(str).unique()) if "Team" in df.columns else [])
            selected_team = st.selectbox("🏈 Team", team_options)
        with filter_col4:
            site_filter = st.selectbox("💰 Site", ["Both", "DK", "FD"])

        sheet_cols = list(df.columns)
        key_id = f"{selected_sport}::{selected_dataset}::{selected_sheet}"
        sname = selected_sheet.strip().lower()

        # Default visible columns per sport/sheet
        if selected_sport == "MLB":
            if "pitch" in sname:
                default_visible = [c for c in DESIRED_MLB_PITCHERS if c in sheet_cols]
            elif "batter" in sname or "hit" in sname:
                default_visible = [c for c in DESIRED_MLB_BATTERS if c in sheet_cols]
            elif "stack" in sname:
                default_visible = [c for c in DESIRED_MLB_STACKS if c in sheet_cols]
            else:
                default_visible = sheet_cols
        elif selected_sport == "NASCAR" and ("betting" in sname or "dashboard" in sname):
            default_visible = [c for c in DESIRED_NASCAR_BETTING if c in sheet_cols]
        elif selected_sport == "NASCAR" and "proj" in sname:
            default_visible = [c for c in DESIRED_NASCAR_PROJECTIONS if c in sheet_cols]
        else:
            default_visible = st.session_state.visible_cols.get(key_id, sheet_cols)

        preset_area = st.container()
        with preset_area:
            cols_p = st.columns([1.2, 1.2, 1, 1.2, 2])
            preset_dict = st.session_state.presets.get(key_id, {})
            preset_names = ["(None)"] + sorted(preset_dict.keys())
            selected_preset_name = cols_p[0].selectbox("Presets", preset_names, index=0)
            preset_new_name = cols_p[1].text_input("Save as…", placeholder="e.g., MLB Compact")
            save_btn = cols_p[2].button("Save/Update")
            delete_btn = cols_p[3].button("Delete")
            apply_on_select = cols_p[4].checkbox("Auto-apply on select", value=True)

        options_cols = [c for c in default_visible if c in sheet_cols] if (selected_sport in ["NASCAR","MLB"]) else sheet_cols
        visible_columns = st.multiselect(
            "👁️ Visible Columns",
            options=options_cols,
            default=options_cols,
            help="Select which columns to display",
        )

        reset_clicked = st.button("↩️ Reset to all columns", help="Show every whitelisted column")
        if reset_clicked:
            visible_columns = options_cols[:]
            st.session_state.visible_cols[key_id] = visible_columns

        if selected_preset_name != "(None)" and apply_on_select and not reset_clicked:
            preset_cols = [c for c in preset_dict.get(selected_preset_name, []) if c in options_cols]
            if preset_cols:
                visible_columns = preset_cols

        if save_btn and preset_new_name.strip():
            preset_cols = visible_columns or options_cols
            st.session_state.presets.setdefault(key_id, {})[preset_new_name.strip()] = preset_cols
            _save_preset_store(st.session_state.presets)
            st.success(f"Saved preset: {preset_new_name.strip()}")

        if delete_btn and selected_preset_name != "(None)":
            st.session_state.presets.get(key_id, {}).pop(selected_preset_name, None)
            _save_preset_store(st.session_state.presets)
            st.success(f"Deleted preset: {selected_preset_name}")

        st.session_state.visible_cols[key_id] = visible_columns or options_cols

        # Apply filters
        filtered_df = df.copy()
        name_cols = [c for c in ["Driver", "Player Name"] if c in filtered_df.columns]
        if search_query and name_cols:
            ncol = "Driver" if "Driver" in name_cols else name_cols[0]
            filtered_df = filtered_df[
                filtered_df[ncol].astype(str).str.contains(search_query, case=False, na=False)
            ]
        if selected_pos != "All" and "Pos" in filtered_df.columns:
            filtered_df = filtered_df[filtered_df["Pos"].astype(str) == selected_pos]
        if selected_team != "All" and "Team" in filtered_df.columns:
            filtered_df = filtered_df[filtered_df["Team"].astype(str) == selected_team]
        if min_sal is not None and "DK Sal" in filtered_df.columns and filtered_df["DK Sal"].notna().any():
            ds = pd.to_numeric(filtered_df["DK Sal"], errors="coerce")
            filtered_df = filtered_df[(ds >= min_sal) & (ds <= max_sal)]
        if min_proj is not None and "DK Proj" in filtered_df.columns and filtered_df["DK Proj"].notna().any():
            dp = pd.to_numeric(filtered_df["DK Proj"], errors="coerce")
            filtered_df = filtered_df[(dp >= min_proj) & (dp <= max_proj)]

        # MLB-specific row filters
        if selected_sport == "MLB":
            if bat_min is not None and "Bat Order" in filtered_df.columns:
                bo = pd.to_numeric(filtered_df["Bat Order"], errors="coerce")
                filtered_df = filtered_df[(bo >= bat_min) & (bo <= bat_max)]
            if selected_bats != "All" and "Bats" in filtered_df.columns:
                filtered_df = filtered_df[filtered_df["Bats"].astype(str) == selected_bats]
            if selected_pitch_hand != "All" and "Pitcher Hand" in filtered_df.columns:
                filtered_df = filtered_df[filtered_df["Pitcher Hand"].astype(str) == selected_pitch_hand]
            if k_min is not None and "K Proj" in filtered_df.columns:
                kp = pd.to_numeric(filtered_df["K Proj"], errors="coerce")
                filtered_df = filtered_df[(kp >= k_min) & (kp <= k_max)]
            if ip_min is not None and "IP Proj" in filtered_df.columns:
                ip = pd.to_numeric(filtered_df["IP Proj"], errors="coerce")
                filtered_df = filtered_df[(ip >= ip_min) & (ip <= ip_max)]

        # Site filtering
        def filter_site(df_in: pd.DataFrame, site: str) -> pd.DataFrame:
            if site == "Both":
                return df_in
            drop_prefix = "FD " if site == "DK" else "DK "
            drop_cols = [c for c in df_in.columns if c.startswith(drop_prefix)]
            return df_in.drop(columns=drop_cols, errors="ignore")

        pruned_df = filter_site(filtered_df, site_filter)

        final_cols = [c for c in (visible_columns or options_cols) if c in pruned_df.columns]
        if "Driver" in final_cols:
            final_cols = ["Driver"] + [c for c in final_cols if c != "Driver"]

        display_df = pruned_df[final_cols] if final_cols else pruned_df
        display_df = _round_numeric_for_display(display_df)

# ===============================
# SAFE TABLE RENDER + CLEAN EXPORT
# ===============================
base_df = pruned_df if "pruned_df" in locals() else df

if base_df is None or not isinstance(base_df, pd.DataFrame) or base_df.empty:
    st.warning("No data to display after filters.")
else:
    try:
        final_cols = [c for c in (visible_columns or options_cols) if c in base_df.columns]
    except Exception:
        final_cols = list(base_df.columns)

    if "Driver" in final_cols:
        final_cols = ["Driver"] + [c for c in final_cols if c != "Driver"]

    display_df = base_df[final_cols] if final_cols else base_df

    # Count BEFORE any Streamlit display calls
    total_rows = len(df) if isinstance(df, pd.DataFrame) else 0
    shown_rows = len(display_df) if isinstance(display_df, pd.DataFrame) else 0
    st.markdown(f"**Showing {shown_rows} of {total_rows} rows**")

# keep numbers numeric for correct sorting
display_df_clean = _round_numeric_for_display(display_df)

# normalize salary columns to integers for display/export
for col in ["DK Sal", "FD Sal", "Stack Salary"]:
    if col in display_df_clean.columns:
        s = pd.to_numeric(display_df_clean[col], errors="coerce")
        display_df_clean[col] = s.round(0).astype("Int64")  # plain integers, keeps NA support

# force certain metrics to 1-decimal numeric (handles object dtypes like strings)
ONE_DEC_COLS = ["pFL", "pLL", "DK PP", "FD PP", "DK Dom", "FD Dom", "Proj Fin"]
for col in ONE_DEC_COLS:
    if col in display_df_clean.columns:
        display_df_clean[col] = pd.to_numeric(display_df_clean[col], errors="coerce").round(1)

st.dataframe(
    display_df_clean,
    use_container_width=True,
    height=420,
    column_config=build_column_config(display_df_clean),
)

# Export: numeric CSV (no $, %, etc.)
try:
    export_filename = f"{selected_sport}_{selected_dataset}_{selected_sheet}_filtered.csv".replace(" ", "_")
    export_bytes = display_df_clean.to_csv(index=False).encode("utf-8")
    st.download_button(
        "📥 Export Filtered Data",
        data=export_bytes,
        file_name=export_filename,
        mime="text/csv",
        help="Download the filtered data as CSV",
    )
except Exception as e:
    st.warning(f"Could not generate export (CSV build failed): {e}")

with tab2:
    st.subheader("📊 Advanced Analytics")
    chart_df = pruned_df if "pruned_df" in locals() else df
    render_analytics_auto(chart_df, selected_sport, selected_sheet, site_filter)

with tab3:
    if is_nfl_projections_context(selected_sport, selected_sheet, df) and "Pos" in df.columns:
        st.subheader("📋 Position Summary (NFL Projections)")
        sites = ["DK", "FD"] if site_filter == "Both" else ([site_filter] if site_filter in ["DK","FD"] else ["DK","FD"])
        for site in sites:
            sm = nfl_position_summary(pruned_df if "pruned_df" in locals() else df, site)
            if sm.empty:
                continue
            st.markdown(f"**{site} Summary**")
            st.dataframe(sm, use_container_width=True, column_config=build_column_config(sm))
    else:
        st.info("Position Summary is available only for NFL Projections sheets.")
